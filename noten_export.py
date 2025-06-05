from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from daten_laden import lade_noten

def noten_exportieren():
    # Dateispeicherdialog öffnen
    pfad = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel-Dateien", "*.xlsx")],
        title="Noten exportieren als Excel"
    )

    if not pfad:
        return

    try:
        # Noten aus der Datenbank laden
        daten = lade_noten()

        wb = Workbook()
        ws = wb.active
        ws.title = "Notenübersicht"

        # Spaltenüberschriften
        header = [
            "Schüler", "Geschlecht", "Ort", "Postleitzahl", "Klasse", "Eintrittsjahr",
            "Fach", "Note", "Notenart", "Datum", "Lehrer", "Status"
        ]
        ws.append(header)

        # Kopfzeilen-Stil definieren
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Kopfzeilen-Stil anwenden
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(header)):
            for cell in col:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border

        # Farbliche Füllung je nach Status definieren
        farben = {
            "Nicht gefährdet": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "Beobachten": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "Gefährdet": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
            "Offen": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        }

        # Notendaten einfügen und einfärben
        for zeile in daten:
            try:
                note = int(zeile[7])
            except Exception:
                note = None

            # Status berechnen
            if note is None:
                status = "Offen"
            elif note <= 3:
                status = "Nicht gefährdet"
            elif note == 4:
                status = "Beobachten"
            else:
                status = "Gefährdet"

            # Datum formatieren
            try:
                datum_obj = zeile[9]
                datum = datum_obj.strftime("%d.%m.%Y")
            except Exception:
                datum = zeile[9]

            # Neue Datenzeile einfügen
            daten_zeile = [
                zeile[0],  # Schüler
                zeile[1],  # Geschlecht
                zeile[2],  # Ort
                zeile[3],  # Postleitzahl
                zeile[4],  # Klasse
                zeile[5],  # Eintrittsjahr
                zeile[6],  # Fach
                zeile[7],  # Note
                zeile[8],  # Notenart
                datum,     # Datum
                zeile[10], # Lehrer
                status     # Status
            ]
            ws.append(daten_zeile)

            # Letzte Zeile abrufen
            letzte_zeile = ws.max_row
            # Farbfüllung je nach Status anwenden
            for col in range(1, len(daten_zeile)+1):
                zelle = ws.cell(row=letzte_zeile, column=col)
                zelle.fill = farben[status]
                zelle.border = border
                zelle.alignment = Alignment(horizontal="center")

        # Spaltenbreite automatisch anpassen
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2

        # Datei speichern
        wb.save(pfad)
        messagebox.showinfo("Export erfolgreich", f"Excel-Datei wurde gespeichert:\n{pfad}")

    except Exception as e:
        # Fehlermeldung bei Problemen
        messagebox.showerror("Fehler beim Exportieren", str(e))
